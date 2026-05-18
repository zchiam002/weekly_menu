"""Build dish-rotation.xlsx — cook log + dish rotation tracker."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = r"D:\ZLStuff\development\weekly_menus\dish-rotation.xlsx"

wb = Workbook()

header_fill = PatternFill("solid", start_color="305496")
header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
data_font = Font(name="Arial", size=10)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ============ Sheet: Cook log ============
log = wb.active
log.title = "Cook log"

log_headers = ["Date", "Day", "Dish", "Notes"]
for i, h in enumerate(log_headers, 1):
    c = log.cell(row=1, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

log_widths = [14, 8, 42, 32]
for i, w in enumerate(log_widths, 1):
    log.column_dimensions[get_column_letter(i)].width = w
log.row_dimensions[1].height = 22
log.freeze_panes = "A2"

# Day column auto-derived from Date
for r in range(2, 301):
    log.cell(row=r, column=2, value=f'=IF(A{r}="","",TEXT(A{r},"ddd"))').font = data_font
    log.cell(row=r, column=1).number_format = "yyyy-mm-dd"

log.auto_filter.ref = "A1:D300"

# ============ Sheet: Dish rotation ============
rot = wb.create_sheet("Dish rotation")

rot_headers = ["Dish", "Category", "Last cooked", "Days since",
               "Cook count", "Cool-down (days)", "Status", "Notes"]
for i, h in enumerate(rot_headers, 1):
    c = rot.cell(row=1, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

rot_widths = [40, 14, 14, 12, 12, 16, 16, 32]
for i, w in enumerate(rot_widths, 1):
    rot.column_dimensions[get_column_letter(i)].width = w
rot.row_dimensions[1].height = 22
rot.freeze_panes = "A2"

# (Dish, Category, Cool-down days, Notes)
dishes = [
    # This week (May 11-15, 2026)
    ("Taiwanese pork chop", "Pork", 21, ""),
    ("Fried rice with prawns", "Prawn", 14, ""),
    ("Steamed pomfret", "Fish", 21, ""),
    ("Pork rib king", "Pork", 21, ""),
    ("Eggplant with minced pork", "Pork", 21, ""),
    ("Ee-fu mee", "Noodle", 21, ""),
    ("HK style steamed seabass", "Fish", 21, ""),
    ("Steamed egg (ah gong style)", "Egg", 14, ""),
    ("Radish pork rib soup with dates", "Soup", 21, ""),
    ("Wasabi mayo prawns", "Prawn", 21, ""),
    ("Chicken chop", "Chicken", 21, ""),
    ("Mash potatoes", "Side", 21, ""),
    ("Roasted vegetables (oven)", "Vegetable", 21, ""),

    # Pork
    ("Sticky pork belly", "Pork", 21, ""),
    ("Taiwanese braised pork rice", "Pork", 28, ""),
    ("Steamed fish paste with minced meat", "Fish", 28, ""),
    ("Steamed tofu with minced pork", "Egg", 21, ""),

    # Chicken
    ("Ah gong gai fan", "Chicken", 21, ""),
    ("Sesame chicken mee sua", "Chicken", 21, ""),
    ("Teriyaki chicken cutlet", "Chicken", 28, ""),
    ("Macau Portuguese chicken rice", "Chicken", 35, ""),
    ("Claypot chicken rice", "Chicken", 28, ""),
    ("Baked chicken wings", "Chicken", 21, ""),

    # Fish
    ("Steamed baby threadfin", "Fish", 21, ""),
    ("Salmon kabayaki", "Fish", 28, ""),
    ("Lemon butter salmon", "Fish", 28, ""),
    ("Assam fish", "Fish", 28, ""),
    ("Thai style steamed seabass", "Fish", 28, ""),
    ("Fish soup", "Fish", 21, ""),

    # Prawn / Seafood
    ("Prawn omelette", "Prawn", 21, ""),

    # Rice
    ("Pineapple fried rice", "Rice", 28, ""),
    ("Egg fried rice", "Rice", 14, ""),
    ("Yam rice", "Rice", 28, ""),
    ("Hokkien kiam png", "Rice", 28, ""),

    # Noodle
    ("Stir fried bee hoon", "Noodle", 14, ""),
    ("Fried udon", "Noodle", 21, ""),
    ("Hong Kong style fried noodles", "Noodle", 21, ""),
    ("Pork noodle soup", "Noodle", 21, ""),
    ("Ah gong fried prawn noodle", "Noodle", 28, ""),

    # Soup
    ("ABC soup", "Soup", 14, ""),
    ("Watercress pork rib soup", "Soup", 21, ""),
    ("Old cucumber pork rib soup", "Soup", 21, ""),
    ("Winter melon pork rib soup", "Soup", 28, ""),
    ("Pork rib spinach soup", "Soup", 21, ""),
    ("Mushroom soup", "Soup", 28, ""),
    ("Fishball soup with cabbage", "Soup", 21, ""),

    # Egg / Tofu
    ("Tomato egg", "Egg", 14, ""),
    ("Egg with preserved radish", "Egg", 21, ""),
    ("Three egg spinach broth", "Egg", 28, ""),
    ("Steamed tofu with soy garlic sauce", "Egg", 21, ""),

    # Vegetable
    ("Stir fried Sichuan string beans", "Vegetable", 21, ""),
    ("Stir fried baby spinach", "Vegetable", 14, ""),
    ("Broccoli stir fry with mushroom", "Vegetable", 14, ""),
    ("Tofu hotplate with broccoli", "Vegetable", 21, ""),
    ("Moroccan vegetable stew", "Vegetable", 28, ""),

    # Sides / Other
    ("Hey zho (prawn rolls)", "Side", 21, ""),
    ("Thai fishcakes", "Side", 28, ""),
    ("Beef steak", "Beef", 28, ""),
    ("Cubano sandwich", "Western", 35, ""),
    ("Philly cheesesteak", "Western", 35, ""),
    ("Green curry", "Other", 28, ""),
]

LOG_RANGE_DATE = "'Cook log'!A2:A1000"
LOG_RANGE_DISH = "'Cook log'!C2:C1000"

for idx, (dish, cat, cooldown, notes) in enumerate(dishes, 2):
    r = idx
    rot.cell(row=r, column=1, value=dish).font = data_font
    rot.cell(row=r, column=2, value=cat).font = data_font
    # Last cooked = MAXIFS over Cook log
    rot.cell(row=r, column=3,
        value=f'=IFERROR(IF(MAXIFS({LOG_RANGE_DATE},{LOG_RANGE_DISH},A{r})=0,"",'
              f'MAXIFS({LOG_RANGE_DATE},{LOG_RANGE_DISH},A{r})),"")'
    ).font = data_font
    rot.cell(row=r, column=3).number_format = "yyyy-mm-dd"
    # Days since today
    rot.cell(row=r, column=4,
        value=f'=IF(C{r}="","",TODAY()-C{r})'
    ).font = data_font
    # Cook count
    rot.cell(row=r, column=5,
        value=f'=COUNTIF({LOG_RANGE_DISH},A{r})'
    ).font = data_font
    # Cool-down target
    rot.cell(row=r, column=6, value=cooldown).font = data_font
    rot.cell(row=r, column=6).alignment = Alignment(horizontal="center", vertical="center")
    # Status
    rot.cell(row=r, column=7,
        value=f'=IF(C{r}="","Never cooked",IF(D{r}<F{r},"Cooldown","Due"))'
    ).font = data_font
    rot.cell(row=r, column=7).alignment = Alignment(horizontal="center", vertical="center")
    # Notes
    rot.cell(row=r, column=8, value=notes).font = data_font

    for c in range(1, 9):
        rot.cell(row=r, column=c).border = border
        if c not in (6, 7):
            rot.cell(row=r, column=c).alignment = Alignment(vertical="center")

last_row = len(dishes) + 1
rot.auto_filter.ref = f"A1:H{last_row}"

# Conditional formatting on Status column G
green_fill = PatternFill("solid", start_color="C6EFCE")
yellow_fill = PatternFill("solid", start_color="FFEB9C")
blue_fill = PatternFill("solid", start_color="D9E1F2")

green_font_c = Font(name="Arial", size=10, color="006100", bold=True)
yellow_font_c = Font(name="Arial", size=10, color="9C5700")
blue_font_c = Font(name="Arial", size=10, color="2F5496", italic=True)

rot.conditional_formatting.add(f"G2:G{last_row}",
    CellIsRule(operator="equal", formula=['"Due"'], fill=green_fill, font=green_font_c))
rot.conditional_formatting.add(f"G2:G{last_row}",
    CellIsRule(operator="equal", formula=['"Cooldown"'], fill=yellow_fill, font=yellow_font_c))
rot.conditional_formatting.add(f"G2:G{last_row}",
    CellIsRule(operator="equal", formula=['"Never cooked"'], fill=blue_fill, font=blue_font_c))

# ============ Sheet: Reference ============
ref = wb.create_sheet("Reference")
ref.cell(row=1, column=1, value="Categories").font = Font(bold=True, name="Arial", size=11)

categories = sorted(set(d[1] for d in dishes))
for i, c in enumerate(categories, 2):
    ref.cell(row=i, column=1, value=c)
ref.column_dimensions["A"].width = 18

# Data validation: dish dropdown on Cook log column C (warning, not stop — allow free entries)
dv_dish = DataValidation(
    type="list",
    formula1=f"='Dish rotation'!$A$2:$A${len(dishes)+1}",
    allow_blank=True,
    errorStyle="warning",
    error="Dish not in master list. Add it to the Dish rotation tab to start tracking it.",
    errorTitle="New dish?",
)
dv_dish.add("C2:C300")
log.add_data_validation(dv_dish)

# Category dropdown on Dish rotation column B
dv_cat = DataValidation(
    type="list",
    formula1=f"=Reference!$A$2:$A${len(categories)+1}",
    allow_blank=True,
)
dv_cat.add("B2:B500")
rot.add_data_validation(dv_cat)

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Dishes: {len(dishes)}")
print(f"Categories: {len(categories)}  ->  {categories}")

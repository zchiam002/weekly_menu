"""Build kitchen-inventory.xlsx for D:\\ZLStuff\\development\\weekly_menus."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = r"D:\ZLStuff\development\weekly_menus\kitchen-inventory.xlsx"

wb = Workbook()

# ============ Sheet: Inventory ============
inv = wb.active
inv.title = "Inventory"

headers = ["Item", "Category", "Status", "Quantity", "Location", "Expiry", "Last updated", "Notes"]
header_fill = PatternFill("solid", start_color="305496")
header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
data_font = Font(name="Arial", size=10)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for i, h in enumerate(headers, 1):
    c = inv.cell(row=1, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

widths = [32, 24, 12, 14, 18, 12, 14, 36]
for i, w in enumerate(widths, 1):
    inv.column_dimensions[get_column_letter(i)].width = w
inv.row_dimensions[1].height = 22

# Items: (Item, Category, Status, Quantity, Location, Expiry, Last updated, Notes)
items = [
    # Sauces & Condiments
    ("Light soy sauce", "Sauces & Condiments", "In stock", "", "Pantry", "", "", ""),
    ("Dark soy sauce (premium)", "Sauces & Condiments", "In stock", "", "Pantry", "", "", "For braising / dark colour"),
    ("Oyster sauce", "Sauces & Condiments", "In stock", "", "Pantry", "", "", ""),
    ("Sesame oil", "Sauces & Condiments", "In stock", "", "Pantry", "", "", ""),
    ("Shaoxing Huatiao wine", "Sauces & Condiments", "In stock", "", "Pantry", "", "", ""),
    ("Plum sauce", "Sauces & Condiments", "In stock", "", "Pantry", "", "", "Pork rib king gravy"),
    ("A1 sauce", "Sauces & Condiments", "In stock", "", "Pantry", "", "", "Pork rib king gravy"),
    ("HP sauce", "Sauces & Condiments", "In stock", "", "Pantry", "", "", "Pork rib king gravy"),
    ("Chilli sauce", "Sauces & Condiments", "In stock", "", "Pantry", "", "", ""),
    ("Ketchup", "Sauces & Condiments", "In stock", "", "Pantry", "", "", ""),
    ("Spicy broad bean paste 辣豆瓣酱", "Sauces & Condiments", "In stock", "", "Pantry", "", "", "Eggplant w/ minced pork"),
    ("Black vinegar", "Sauces & Condiments", "In stock", "", "Pantry", "", "", ""),
    ("White vinegar", "Sauces & Condiments", "In stock", "", "Pantry", "", "", "For mayo"),
    ("Mustard", "Sauces & Condiments", "In stock", "", "Pantry", "", "", "For mayo"),
    ("Balsamic vinegar", "Sauces & Condiments", "In stock", "", "Pantry", "", "", ""),
    ("Olive oil (extra virgin)", "Sauces & Condiments", "In stock", "", "Pantry", "", "", ""),
    ("Sunflower / neutral oil", "Sauces & Condiments", "In stock", "", "Pantry", "", "", ""),
    ("Kewpie mayonnaise", "Sauces & Condiments", "In stock", "", "Fridge - door", "", "", "For wasabi mayo prawns"),
    ("Wasabi paste / tube", "Sauces & Condiments", "In stock", "", "Fridge - door", "", "", ""),
    ("Kicap manis", "Sauces & Condiments", "In stock", "", "Pantry", "", "", ""),
    ("Fish sauce", "Sauces & Condiments", "In stock", "", "Pantry", "", "", ""),
    ("Orange marmalade", "Sauces & Condiments", "In stock", "", "Pantry", "", "", "Optional - orange mayo"),

    # Spices & Seasonings
    ("Salt", "Spices & Seasonings", "In stock", "", "Spice rack", "", "", ""),
    ("Sugar (white)", "Spices & Seasonings", "In stock", "", "Pantry", "", "", ""),
    ("Brown sugar", "Spices & Seasonings", "In stock", "", "Pantry", "", "", ""),
    ("White pepper", "Spices & Seasonings", "In stock", "", "Spice rack", "", "", ""),
    ("Black pepper (ground)", "Spices & Seasonings", "In stock", "", "Spice rack", "", "", ""),
    ("Five spice powder", "Spices & Seasonings", "In stock", "", "Spice rack", "", "", ""),
    ("Curry powder (meat)", "Spices & Seasonings", "In stock", "", "Spice rack", "", "", ""),
    ("Chicken stock powder", "Spices & Seasonings", "In stock", "", "Spice rack", "", "", ""),
    ("Chicken stock cubes", "Spices & Seasonings", "In stock", "", "Pantry", "", "", ""),
    ("Baking soda", "Spices & Seasonings", "In stock", "", "Pantry", "", "", "Meat tenderiser"),
    ("Cornflour / cornstarch", "Spices & Seasonings", "In stock", "", "Pantry", "", "", ""),
    ("All-purpose flour", "Spices & Seasonings", "In stock", "", "Pantry", "", "", ""),
    ("Tapioca flour", "Spices & Seasonings", "In stock", "", "Pantry", "", "", ""),
    ("Garlic powder", "Spices & Seasonings", "In stock", "", "Spice rack", "", "", ""),
    ("Smoked paprika", "Spices & Seasonings", "In stock", "", "Spice rack", "", "", ""),
    ("Herbs de Provence", "Spices & Seasonings", "In stock", "", "Spice rack", "", "", ""),
    ("Ground cumin", "Spices & Seasonings", "In stock", "", "Spice rack", "", "", ""),
    ("Ground turmeric", "Spices & Seasonings", "In stock", "", "Spice rack", "", "", ""),
    ("Ground cinnamon", "Spices & Seasonings", "In stock", "", "Spice rack", "", "", ""),

    # Dry Goods
    ("Jasmine rice", "Dry Goods", "In stock", "", "Pantry", "", "", ""),
    ("Bee hoon (rice vermicelli)", "Dry Goods", "In stock", "", "Pantry", "", "", ""),
    ("Ee-fu noodles", "Dry Goods", "In stock", "", "Pantry", "", "", ""),
    ("Mee sua", "Dry Goods", "In stock", "", "Pantry", "", "", ""),
    ("Hong Kong egg noodles", "Dry Goods", "In stock", "", "Pantry", "", "", ""),
    ("Pasta", "Dry Goods", "In stock", "", "Pantry", "", "", ""),
    ("Dried Chinese mushrooms (shiitake)", "Dry Goods", "In stock", "", "Pantry", "", "", ""),
    ("Red dates / jujubes", "Dry Goods", "In stock", "", "Pantry", "", "", "Soup ingredient"),
    ("Goji berries (枸杞)", "Dry Goods", "In stock", "", "Pantry", "", "", "Optional, soups"),
    ("Cashew nuts", "Dry Goods", "In stock", "", "Pantry", "", "", ""),
    ("Raisins", "Dry Goods", "In stock", "", "Pantry", "", "", "Pineapple rice"),
    ("Pork floss", "Dry Goods", "In stock", "", "Pantry", "", "", ""),

    # Canned/Jarred
    ("Coconut milk (can)", "Canned/Jarred", "In stock", "", "Pantry", "", "", ""),
    ("Evaporated milk (can)", "Canned/Jarred", "In stock", "", "Pantry", "", "", ""),
    ("Diced tomatoes (can)", "Canned/Jarred", "In stock", "", "Pantry", "", "", ""),
    ("Tomato paste", "Canned/Jarred", "In stock", "", "Pantry", "", "", ""),
    ("Chickpeas (can)", "Canned/Jarred", "In stock", "", "Pantry", "", "", "Moroccan stew"),
    ("Canned mushrooms (straw)", "Canned/Jarred", "In stock", "", "Pantry", "", "", ""),
    ("Bamboo shoots (can)", "Canned/Jarred", "In stock", "", "Pantry", "", "", ""),
    ("Chicken stock (carton)", "Canned/Jarred", "In stock", "", "Pantry", "", "", ""),

    # Fridge - Dairy & Eggs
    ("Eggs", "Fridge - Dairy & Eggs", "In stock", "", "Fridge - door", "", "", ""),
    ("Pasteurised eggs", "Fridge - Dairy & Eggs", "In stock", "", "Fridge - door", "", "", ""),
    ("Butter (unsalted)", "Fridge - Dairy & Eggs", "In stock", "", "Fridge - door", "", "", ""),
    ("Milk", "Fridge - Dairy & Eggs", "In stock", "", "Fridge - door", "", "", ""),
    ("Heavy cream", "Fridge - Dairy & Eggs", "In stock", "", "Fridge - door", "", "", ""),
    ("Mozzarella", "Fridge - Dairy & Eggs", "In stock", "", "Fridge - top", "", "", ""),
    ("Cheddar", "Fridge - Dairy & Eggs", "In stock", "", "Fridge - top", "", "", ""),
    ("Parmesan", "Fridge - Dairy & Eggs", "In stock", "", "Fridge - top", "", "", ""),

    # Freezer - Proteins
    ("Pork chops / loin", "Freezer - Proteins", "In stock", "", "Freezer", "", "", ""),
    ("Pork ribs", "Freezer - Proteins", "In stock", "", "Freezer", "", "", ""),
    ("Pork belly", "Freezer - Proteins", "In stock", "", "Freezer", "", "", ""),
    ("Pork lard", "Freezer - Proteins", "In stock", "", "Freezer", "", "", ""),
    ("Minced pork", "Freezer - Proteins", "In stock", "", "Freezer", "", "", ""),
    ("Sliced pork", "Freezer - Proteins", "In stock", "", "Freezer", "", "", ""),
    ("Chicken thighs (boneless)", "Freezer - Proteins", "In stock", "", "Freezer", "", "", ""),
    ("Chicken wings", "Freezer - Proteins", "In stock", "", "Freezer", "", "", ""),
    ("Whole chicken / pieces", "Freezer - Proteins", "In stock", "", "Freezer", "", "", ""),
    ("Prawns", "Freezer - Proteins", "In stock", "", "Freezer", "", "", ""),
    ("Salmon", "Freezer - Proteins", "In stock", "", "Freezer", "", "", ""),
    ("Sotong (squid)", "Freezer - Proteins", "In stock", "", "Freezer", "", "", ""),
    ("Fish paste", "Freezer - Proteins", "In stock", "", "Freezer", "", "", ""),
    ("Fishball (Fuzhou)", "Freezer - Proteins", "In stock", "", "Freezer", "", "", ""),
    ("Hey zho (prawn rolls)", "Freezer - Proteins", "In stock", "", "Freezer", "", "", ""),
    ("Roast pork (sio bak)", "Freezer - Proteins", "In stock", "", "Freezer", "", "", ""),

    # Produce
    ("Onions", "Produce", "In stock", "", "Pantry", "", "", ""),
    ("Red onion", "Produce", "In stock", "", "Pantry", "", "", ""),
    ("Garlic", "Produce", "In stock", "", "Pantry", "", "", ""),
    ("Ginger (young)", "Produce", "In stock", "", "Fridge - crisper", "", "", ""),
    ("Spring onion", "Produce", "In stock", "", "Fridge - crisper", "", "", ""),
    ("Coriander", "Produce", "In stock", "", "Fridge - crisper", "", "", ""),
    ("Lemon", "Produce", "In stock", "", "Fridge - crisper", "", "", ""),
    ("Lime", "Produce", "In stock", "", "Fridge - crisper", "", "", ""),
    ("Carrots", "Produce", "In stock", "", "Fridge - crisper", "", "", ""),
    ("Potatoes", "Produce", "In stock", "", "Pantry", "", "", ""),
    ("Sweet potato", "Produce", "In stock", "", "Pantry", "", "", ""),
    ("Tomatoes", "Produce", "In stock", "", "Fridge - crisper", "", "", ""),
    ("Birds eye chilli", "Produce", "In stock", "", "Fridge - crisper", "", "", ""),
    ("Red chilli", "Produce", "In stock", "", "Fridge - crisper", "", "", ""),
    ("Green chilli", "Produce", "In stock", "", "Fridge - crisper", "", "", ""),

    # Specialty / Asian
    ("Tofu (silken)", "Specialty / Asian", "In stock", "", "Fridge - top", "", "", ""),
    ("Egg tofu", "Specialty / Asian", "In stock", "", "Fridge - top", "", "", ""),
    ("Tow pok (fried tofu puffs)", "Specialty / Asian", "In stock", "", "Pantry", "", "", ""),
    ("Beancurd skin", "Specialty / Asian", "In stock", "", "Pantry", "", "", ""),
    ("Laksa paste", "Specialty / Asian", "In stock", "", "Pantry", "", "", ""),
    ("Pork rib soup paste (bak kut teh)", "Specialty / Asian", "In stock", "", "Pantry", "", "", ""),
    ("Green curry paste", "Specialty / Asian", "In stock", "", "Pantry", "", "", ""),
    ("Red curry paste", "Specialty / Asian", "In stock", "", "Pantry", "", "", ""),
    ("Dashi", "Specialty / Asian", "In stock", "", "Pantry", "", "", ""),
    ("Tobiko", "Specialty / Asian", "In stock", "", "Freezer", "", "", ""),
    ("Salted egg", "Specialty / Asian", "In stock", "", "Fridge - top", "", "", ""),
    ("Century egg", "Specialty / Asian", "In stock", "", "Pantry", "", "", ""),
    ("Mui Heong salted fish", "Specialty / Asian", "In stock", "", "Freezer", "", "", ""),
    ("Preserved radish (chai poh)", "Specialty / Asian", "In stock", "", "Pantry", "", "", ""),
]

for row_idx, item in enumerate(items, 2):
    for col_idx, val in enumerate(item, 1):
        c = inv.cell(row=row_idx, column=col_idx, value=val)
        c.font = data_font
        c.border = border
        c.alignment = Alignment(vertical="center")

inv.freeze_panes = "A2"
inv.auto_filter.ref = f"A1:H{len(items)+1}"

# ============ Sheet: Reference (created early so dropdowns can point to it) ============
ref = wb.create_sheet("Reference")
categories = sorted({it[1] for it in items})
statuses = ["In stock", "Low", "Out"]
locations = sorted({it[4] for it in items})

ref.cell(row=1, column=1, value="Categories").font = Font(bold=True, name="Arial", size=11)
ref.cell(row=1, column=2, value="Status").font = Font(bold=True, name="Arial", size=11)
ref.cell(row=1, column=3, value="Locations").font = Font(bold=True, name="Arial", size=11)

for i, c in enumerate(categories, 2):
    ref.cell(row=i, column=1, value=c)
for i, s in enumerate(statuses, 2):
    ref.cell(row=i, column=2, value=s)
for i, l in enumerate(locations, 2):
    ref.cell(row=i, column=3, value=l)

ref.column_dimensions["A"].width = 28
ref.column_dimensions["B"].width = 12
ref.column_dimensions["C"].width = 20

# Data validations on Inventory
dv_cat = DataValidation(type="list", formula1=f"=Reference!$A$2:$A${len(categories)+1}", allow_blank=True)
dv_cat.add("B2:B500")
inv.add_data_validation(dv_cat)

dv_stat = DataValidation(type="list", formula1=f"=Reference!$B$2:$B${len(statuses)+1}", allow_blank=True)
dv_stat.add("C2:C500")
inv.add_data_validation(dv_stat)

dv_loc = DataValidation(type="list", formula1=f"=Reference!$C$2:$C${len(locations)+1}", allow_blank=True)
dv_loc.add("E2:E500")
inv.add_data_validation(dv_loc)

# Conditional formatting on Status column
green_fill = PatternFill("solid", start_color="C6EFCE")
yellow_fill = PatternFill("solid", start_color="FFEB9C")
red_fill = PatternFill("solid", start_color="FFC7CE")
green_font = Font(name="Arial", size=10, color="006100")
yellow_font = Font(name="Arial", size=10, color="9C5700")
red_font = Font(name="Arial", size=10, color="9C0006")

inv.conditional_formatting.add("C2:C500",
    CellIsRule(operator="equal", formula=['"In stock"'], fill=green_fill, font=green_font))
inv.conditional_formatting.add("C2:C500",
    CellIsRule(operator="equal", formula=['"Low"'], fill=yellow_fill, font=yellow_font))
inv.conditional_formatting.add("C2:C500",
    CellIsRule(operator="equal", formula=['"Out"'], fill=red_fill, font=red_font))

# ============ Sheet: Shopping list (insert at index 1, after Inventory) ============
shop = wb.create_sheet("Shopping list", 1)

shop_header_fill = PatternFill("solid", start_color="C00000")
for i, h in enumerate(headers, 1):
    c = shop.cell(row=1, column=i, value=h)
    c.font = header_font
    c.fill = shop_header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

for i, w in enumerate(widths, 1):
    shop.column_dimensions[get_column_letter(i)].width = w
shop.row_dimensions[1].height = 22

# FILTER formula (Excel 365 / 2021+) — pulls Low or Out rows from Inventory
shop["A2"] = ('=FILTER(Inventory!A2:H500, '
              '(Inventory!C2:C500="Low")+(Inventory!C2:C500="Out"), '
              '"All items in stock — nothing to buy!")')
shop["A2"].font = data_font

# Note row
shop.cell(row=25, column=1, value=(
    "Note: list auto-updates from Inventory tab (requires Excel 365/2021+). "
    "On older Excel or LibreOffice, use AutoFilter on the Inventory tab → filter Status to Low/Out."
)).font = Font(name="Arial", size=9, italic=True, color="666666")
shop.merge_cells("A25:H25")

shop.freeze_panes = "A2"

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Items seeded: {len(items)}")
print(f"Categories: {len(categories)}")
print(f"Locations: {len(locations)}")
